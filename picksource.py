#!/usr/local/bin/python
# -*- coding:utf-8 -*-

import sys
import os
import re
import logging
import argparse
import uuid
from openpyxl import Workbook

import constants
import utility 


g_logger = None
def mylogger(path = ''):
    global g_logger
    if g_logger is None:
        g_logger = utility.get_logger(path)

    return g_logger
 

class SourceItem(object):
    uni_ids = set()

    def __init__(self, name, value, res_file, feature, textid = ''):
        """
        name: ID-In-Source in out xlsx
        res_file: Path-of-Source in out xlsx
        textid: TextID in source Table
        value: English-GB in source Table
        """
        self.name = name
        self.value = value
        self.res_file = res_file
        self.feature = feature
        if textid == '':
            self.textid = name + '_' + self.get_uniq_id()
        else:
            self.textid = textid

    def get_uniq_id(self):
        uni_id = str(uuid.uuid4())[:8]
        while uni_id in SourceItem.uni_ids:
            mylogger().warn(uni_id + 'id collision, reproduce')
            uni_id = str(uuid.uuid4())[:8]
        SourceItem.uni_ids.add(uni_id)
        return uni_id


    def __repr__(self):
        return '%s, %s, %s, %s' % (self.textid, self.name, self.value, self.res_file)

class SourceToXlsx(object):
    """
    pick id and values from .property file to xlsx
    """
    def __init__(self, root):
        
        self.source_root = os.path.abspath(root)

        self.wb = Workbook()
        self.ws = self.wb.active
        self.col_map = constants.col_in_xlsx(constants.SOURCES_COLUMNS)
        self.row_num = constants.TITLE_ROW

        self.ws['A1'] = ('Generated from folder %s' %
            os.path.abspath(self.source_root))

        self.add_head_row()
        self.pick_source()

    def pick_source(self):
        picker = SourcePicker(self.source_root)
        res_list = picker.pick()
        self.add_res_content(res_list)
        
    def add_head_row(self):
        for col in constants.SOURCES_COLUMNS:
            idx = self.col_map[col] + str(self.row_num)
            self.ws[idx] = col

        self.row_num += 1

    def add_res_content(self, res_list):
        for res in res_list:
            self.add_res_row(res)

    def add_res_row(self, res):
        idx = self.col_map[constants.TEXTID] + str(self.row_num)
        self.ws[idx] = res.textid

        idx = self.col_map[constants.ENGLISHGB] + str(self.row_num)
        self.ws[idx] = res.value

        idx = self.col_map[constants.SOURCEID] + str(self.row_num)
        self.ws[idx] = res.name
        
        idx = self.col_map[constants.SOURCEPATH] + str(self.row_num)
        self.ws[idx] = res.res_file

        idx = self.col_map[constants.FEATURE] + str(self.row_num)
        self.ws[idx] = res.feature

        self.row_num += 1

    def save(self, filename):
        try:
            if filename.find(':') < 0: # not abs path
                out_folder = os.path.abspath(
                        os.path.join(self.source_root, os.pardir))
                filename = out_folder + '\\' + filename
 
            self.wb.save(filename)
        except Exception as e:
            mylogger().error(str(e))
            mylogger().error('%s might be opened by other application' %
                    filename)

    @classmethod
    def load_source_xlsx(cls, xlsx):
        """
        Load the generated sources.xlsx by class SourceToXlsx
        as a dict {uni_id: sourceitem}
        """
        from openpyxl import load_workbook
        wb = load_workbook(filename = xlsx)
        ws = wb.active
        col_map = constants.col_in_xlsx(constants.SOURCES_COLUMNS)
        read_row = constants.TITLE_ROW + 1
        
        idx = col_map[constants.TEXTID] + str(read_row)
        source_map = {}
        while ws[idx].value is not None:
            name_idx = col_map[constants.SOURCEID] + str(read_row)
            value_idx = col_map[constants.ENGLISHGB] + str(read_row)
            path_idx = col_map[constants.SOURCEPATH] + str(read_row)
            fea_idx = col_map[constants.FEATURE] + str(read_row)

            source_map[ws[idx].value] = (
            SourceItem(ws[name_idx].value,
                    ws[value_idx].value, ws[path_idx].value,
                    ws[fea_idx].value, ws[idx].value) 
            )

            read_row += 1 
            idx = col_map[constants.TEXTID] + str(read_row)

        return source_map


class SourceBasic(object):
    def get_res_pattern(self):
        #to match name = value in file except start with #
        return re.compile(r'^(?!#)\s*([^= ]+)\s*=\s*(.*)\s*')

    def get_comment_pattern(self):
        return re.compile(r'^\s*#.*')


class SourceUpdate(SourceBasic):
    def update_file(self, file_name, res_list):
        """
        change the value in file_name with value in res_list
        """
        #TODO: handle the same file has same ID, e.g. music.properties
        # change res_list to map name:value
        res_map = {}
        for item in res_list:
            res_map[item.name] = item.value

        pair = self.get_res_pattern()

        lines = utility.read_as_list(file_name)
        for index, line in enumerate(lines):
            match = pair.match(line)
            if match:
                name = match.groups()[0].strip()
                value = match.groups()[1].strip()
                replace_after = line.rfind(value)
                if name in res_map: 
                    lines[index] = (line[:replace_after] + res_map[name]
                    + '\n')
                    

        utility.save_list(file_name, lines)


class SourcePicker(SourceBasic):
    @classmethod
    def get_feature(cls, filename):
        parent = os.path.abspath(os.path.join(filename, os.pardir))
        featur_name = ''

        while not os.path.exists(parent + '\\manifest.properties'):
            parent2 = os.path.abspath(os.path.join(parent, os.pardir))
            # if look back reach root of drive and no manifest file
            # use direct parent folder as feature name
            if parent == parent2: 
                parent = os.path.abspath(os.path.join(filename, os.pardir))
                break
            parent = parent2
            
        featur_name = os.path.basename(parent)
        return featur_name

    def __init__(self, root):
        self.root = root

    def pick(self):
        return self.pick_folder(self.root)

    def pick_folder(self, folder):
        res_list = []  #store all matched "name = value"
        
        mylogger().info("Picking folder:" + folder)
        for item in os.listdir(folder):
            subpath = os.path.join(folder, item)
            if os.path.isdir(subpath):
                res_list += self.pick_folder(subpath)
            elif self._is_res_file(subpath):
                feature = self.get_feature(subpath)
                res_list += self.pick_file(subpath, feature)
            else:
                mylogger().debug('Skipped file: ' + subpath)
        return res_list

    def pick_file(self, file_name, feature):
        """
        file_name: the file name of .propery file
        output: [SourceItem] of the property file
        """
        mylogger().info("Picking file:" + file_name)

        pair = self.get_res_pattern()

        lines = utility.read_as_list(file_name)

        relative_path = os.path.relpath(file_name, self.root)
        res_list = []
        mismatchs = []
        for line in lines:
            match = pair.match(line)
            if match:
                res_list.append(SourceItem(match.groups()[0].strip(),
                    match.groups()[1].strip(),
                    relative_path, feature))
            else:
                mismatchs.append(line)

        if len(mismatchs) > 0:
            mylogger().warning('In %s mismatched lines:\n %s' % (file_name,
                ''.join(mismatchs)))

        return res_list

    def _is_res_file(self, file_name):
        name_split = os.path.splitext(os.path.basename(file_name))
        if name_split[0] in ['manifest']:
            return False
        if name_split[1] != '.properties':
            return False
        return True
    

def print_duplicate(res_list):
    """
    findout if there is name duplicated in different files
    """
    from collections import defaultdict
    names_dict = defaultdict(list) 
    for res in res_list:
        names_dict[res.name].append(res)

    for key in names_dict:
        if len(names_dict[key]) > 1:
            temp = names_dict[key]
            value = temp[0].value
            duplicated = str(temp[0])
            for t in temp:
                if t.value != value:
                    duplicated += ('\n' + str(t))
                    
            if duplicated != str(temp[0]):
                print( duplicated)


      
def process(root):
    parent = os.path.abspath(os.path.join(root,
        os.pardir))
    # set current dir to out_folder, so output files will be there
    # print("change out dir to:" + out_folder)
    # os.chdir(out_folder)

    try:
        mylogger(parent).info('Starting processing %s' % root)
        convertor = SourceToXlsx(root)
        convertor.save("sources.xlsx")
    except Exception as e:
        mylogger().error(str(e))
        pass
 
def main():
    parser = argparse.ArgumentParser(description='Pick resource string \
        from given folder')

    parser.add_argument('-i','--in', nargs='+', required=True,
                       help='''Define root folder of resource files. It is mandatory.''')

    parser.add_argument('-o', '--out', nargs='+',
                        help='''The result output folder, including log
                        files''')
    
    args = vars(parser.parse_args())
   
    process(os.path.abspath(args['in'][0]))


if __name__ == '__main__':
    main()

