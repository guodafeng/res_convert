#!/usr/local/bin/python
# -*- coding:utf-8 -*-
import sys
import os
import re
import logging
import argparse
from openpyxl import Workbook
import collections


import constants
import utility 
import picksource
# t2 format v2 
#RefName	ModOP	Info	ZoneType	IsMono	IsUK	IsGSM	IsTradUpdatable	English_US	French_CA	LatinESP	English_CA	CHINE_NEW	Brazilian	Croatian	Czech	French	German	Greek	Hebrew	Hungarian	Polish	Romanian	Spanish	Serbian	Slovak	Slovenia	Italian	Turkish	Dutch	Macedonian	Bulgarian	Catalan	Russian	Malay	Danish	CHINE_OLD	Swedish	Arabic	Vietnamese

T2_COL_NAME = (
        'RefName',
	'ModOP',
	'Info',
	'ZoneType',
	'IsMono',
	'IsUK',
	'IsGSM',
        'IsTradUpdatable', # it is added in v2 table 2017-10-24
	'English_US',
	'French_CA',
	'LatinESP',
	'English_CA',
	'English',
	'CHINE_NEW',
	'Brazilian',
	'Croatian',
	'Czech',
	'French',
	'German',
	'Greek',
	'Hebrew',
	'Hungarian',
	'Polish',
	'Romanian',
	'Spanish',
	'Serbian',
	'Slovak',
	'Slovenia',
	'Italian',
	'Turkish',
	'Dutch',
	'Macedonian',
	'Bulgarian',
	'Catalan',
	'Russian',
	'Malay',
	'Danish',
	'CHINE_OLD',
	'Swedish',
	'Arabic',
	'Vietnamese')
ID_TITLE = 'RefName'
FEATURE_TITLE = 'ModOP'
VALUE_TITLE = 'English_US' 

class SourceItemT2(object):
    def __init__(self,  value, feature, textid, layout = '',
            old_id = '', old_value = ''):
        """
        textid: TextID in source Table, trail uuid to be unique
        value: string in source file
        feature: feature it belongs to
        old_id: old format id with name plus uniq random id
        old_value: old value imported into DB with old_id
        """
        self.old_id = old_id
        self.old_value = old_value
        self.value = value
        self.feature = feature
        self.textid = textid
        self.layout = layout


def load_t2_string_table(xlsx):
    """
    load t2 string table, which contains all languages
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename = xlsx)
    ws = wb.get_sheet_by_name('MESSAGE')
    
    rows = ws.rows
    col_map = {name:index for index, name in enumerate(T2_COL_NAME)}
    t2_map = {}
    for row in rows:
        t2_item = SourceItemT2(row[col_map[VALUE_TITLE]].value,
                row[col_map[FEATURE_TITLE]].value,
                row[col_map[ID_TITLE]].value)
        t2_map[t2_item.textid] = t2_item
        
    t2_map.pop(ID_TITLE,'') #remove the title row in map
    return t2_map

def create_id_mapping(t2_map, source_map):
    """
    map the t2 id with old id in DB
    return: list conatils items which old id is not in t2_map
    """
    not_in_t2 = [] # items of old source but not in t2 source
    changed_in_t2 = []
    for fp_id in source_map:
       source_item = source_map[fp_id]
       new_id = create_new_id(source_item.name, source_item.res_file)
       if new_id not in t2_map:
           not_in_t2.append(source_item)
       else:
           t2_map[new_id].old_id = fp_id
           t2_map[new_id].layout = source_item.layout
           source_item.new_id = new_id
           t2_map[new_id].old_value = source_item.value

           if t2_map[new_id].value != source_item.value:
               changed_in_t2.append(t2_map[new_id])
    
    not_in_t2.sort(key=lambda x:x.res_file)
    changed_in_t2.sort(key=lambda x:x.textid)
    return not_in_t2, changed_in_t2


def create_new_id(name, res_file):
    src_path = os.path.splitext(res_file)[0]
    sep = r'[/\\]'
    parts = re.split(sep, src_path)
    #remove apps
    if parts[0] == 'apps':
        parts = parts[1:]
    parts += (name,)
    return ':'.join(parts)
    

def save_t2_source(t2_items, xlsx, desc = ''):
    col_map = constants.col_in_xlsx(constants.T2_SOURCES_COLUMNS)
    def add_head_row(ws, row_num):
        for col in constants.T2_SOURCES_COLUMNS:
            idx = col_map[col] + str(row_num)
            ws[idx] = col

    def add_res_row(ws, res, row_num):
        idx = col_map[constants.TEXTID] + str(row_num)
        ws[idx] = res.textid

        idx = col_map[constants.ENGLISHGB] + str(row_num)
        ws[idx] = res.value

        idx = col_map[constants.OLDID] + str(row_num)
        ws[idx] = res.old_id
        
        idx = col_map[constants.OLDVALUE] + str(row_num)
        ws[idx] = res.old_value

        idx = col_map[constants.FEATURE] + str(row_num)
        ws[idx] = res.feature
        
        idx = col_map[constants.LAYOUT] + str(row_num)
        ws[idx] = res.layout

    def save_source_xlsx(src_items, xlsx):
        wb = Workbook()
        ws = wb.active
        row_num = constants.TITLE_ROW

        ws['A1'] = desc
        add_head_row(ws, row_num)
        row_num += 1

        for src in src_items:
            add_res_row(ws, src, row_num)
            row_num+=1

        wb.save(xlsx)

    save_source_xlsx(t2_items, xlsx)

def load_t2_source(xlsx):
    from openpyxl import load_workbook
    wb = load_workbook(filename = xlsx)
    ws = wb.active
    col_map = constants.col_in_xlsx(constants.T2_SOURCES_COLUMNS)
    read_row = constants.TITLE_ROW + 1
    
    idx = col_map[constants.TEXTID] + str(read_row)
    old_map = {}
    t2_map = {}
    while ws[idx].value is not None:
        value_idx = col_map[constants.ENGLISHGB] + str(read_row)
        old_idx = col_map[constants.OLDID] + str(read_row)
        oldvalue_idx = col_map[constants.OLDVALUE] + str(read_row)
        fea_idx = col_map[constants.FEATURE] + str(read_row)
        lay_idx = col_map[constants.LAYOUT] + str(read_row)

        t2_item = SourceItemT2(ws[value_idx].value,
                    ws[fea_idx].value,
                    ws[idx].value,
                    ws[lay_idx].value,
                    ws[old_idx].value,
                    ws[oldvalue_idx].value
                    )

        t2_map[ws[idx].value] = t2_item
        if ws[old_idx].value:
            old_map[ws[old_idx].value] = t2_item

        read_row += 1 
        idx = col_map[constants.TEXTID] + str(read_row)

    return t2_map, old_map



def fill_source_with_layout(source_map, src_withlayout):
    for key in source_map:
        if key in src_withlayout:
            source_map[key].layout = src_withlayout[key].layout
            source_map[key].value = src_withlayout[key].value
        else:
            print('no layout info found for' + key)


def map_source_t2(t2_file, old_sources_file, layout_sources_file=''):
    """
    create mapping between old source file and t2 file
    source_map: old format id map
    t2_map: t2 format id map
    """
    source_map = \
    picksource.SourceToXlsx.load_source_xlsx(old_sources_file)
    

    if layout_sources_file:
        src_withlayout = \
        picksource.SourceToXlsx.load_source_xlsx(layout_sources_file)
        fill_source_with_layout(source_map, src_withlayout)

    t2_map = load_t2_string_table(t2_file)
    not_in_t2, changed_in_t2 = create_id_mapping(t2_map, source_map)

    return (t2_map, source_map, not_in_t2, changed_in_t2)


def save_map(t2_map, source_map, not_in_t2, changed_in_t2):
    t2_items = [t2_map[key] for key in sorted(t2_map)]
    save_t2_source(t2_items,'map_source_t2.xlsx', 
    'Generated by scripts, include all strings in t2 given table,\
    mapped old id if exists')

    new_items = [t2_map[key] for key in sorted(t2_map) if not \
            t2_map[key].old_id]
    save_t2_source(new_items,'map_new_added_t2.xlsx',
            'New added strings in t2')
    
    save_t2_source(changed_in_t2,'map_changed_in_t2.xlsx',
            'Strings that changed in t2 given table')
    picksource.SourceToXlsx.save_source_xlsx(not_in_t2,
    'map_not_int_t2.xlsx')


def replace_with_t2id(xliff, old_map):
    pattern = re.compile(r'resname="([^"]+)"')
    lines = utility.read_as_list(xliff)
    out_lines = []
    for line in lines:
        match = pattern.search(line)
        if match:
            old_id = match.groups()[0]
            if old_id in old_map and old_map[old_id].textid:
                new_id = 'resname="%s"' % old_map[old_id].textid
                line = re.sub(pattern, new_id, line)

        out_lines.append(line)
    utility.save_list(xliff, out_lines)


def create_mapped_files():
    (t2_map, source_map, not_in_t2, changed_in_t2) = \
        map_source_t2('string_src/Argon_string.xlsx',
        'string_src/sources-inDB.xlsx','string_src/kaios_old'
        ' id_layout_en-GB.XLSX')
    save_map(t2_map, source_map, not_in_t2, changed_in_t2)

def replace_xliff_folder(folder, old_map):
    def is_xliff(name):
        return name[-4:] == '.xlf'

    folder = os.path.abspath(folder)
    for item in os.listdir(folder):
        subpath = os.path.join(folder, item)
        if os.path.isfile(subpath) and is_xliff(subpath):
            replace_with_t2id(subpath, old_map)

def do_replace_xliff():
    t2_map, old_map = load_t2_source('map_source_t2_v3.xlsx')
    replace_xliff_folder('string_src/xliff/batch1_36lan', old_map)


create_mapped_files()

