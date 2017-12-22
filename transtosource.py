#!/usr/local/bin/python
# -*- coding:utf-8 -*-
import sys
import os
import re
import logging
import argparse
from openpyxl import Workbook
from collections import defaultdict
from openpyxl import load_workbook

import pymssql
import constants
import utility 
import picksource

g_logger = None
def mylogger():
    global g_logger
    if g_logger is None:
        g_logger = get_logger()

    return g_logger
 
class TranslateItem(object):
    def __init__(self, account, project, feature, textid, langcode,
            translation):
        self.account = account
        self.project = project
        self.feature = feature
        self.textid = textid
        self.langcode = langcode
        self.translation = translation

class XlsxToSource(object):
    """
    transfer xlsx back to .property files, group translated different 
    languages to different folder
    """
    def load_translate_xlsx(self, xlsx):
        """
        group the translation in xlsx by langcode
        return dict
        """

        wb = load_workbook(filename = xlsx)
        ws = wb.active
        rows = ws.rows

        #index + 1 because the contents started from 2nd column
        col_map = {name:index+1 for index, name in \
                enumerate(constants.TRANSLATIONS_COLUMNS)}
        
        idx = col_map[constants.TEXTID]
        translations = []

        for num, row in enumerate(rows):
            if num < constants.TITLE_ROW:
                continue

            acc_idx = col_map[constants.ACCOUNT]
            prj_idx = col_map[constants.PROJECT]
            fea_idx = col_map[constants.FEATURE] 
            lan_idx = col_map[constants.LANGCODE]
            tran_idx = col_map[constants.TRANSLATION]

            trans_item = \
            TranslateItem(row[acc_idx].value, row[prj_idx].value, 
                row[fea_idx].value, row[idx].value, row[lan_idx].value, 
                row[tran_idx].value) 

            translations.append(trans_item)
            idx = col_map[constants.TEXTID]

        return translations

    def load_translate_folder(self, folder):
        def is_xlsx(name):
            return name[-5:] == '.xlsx'

        translations = []
        folder = os.path.abspath(folder)
        for item in os.listdir(folder):
            subpath = os.path.join(folder, item)
            if os.path.isfile(subpath) and is_xlsx(subpath):
                translations += self.load_translate_xlsx(subpath)

        return translations




    def _get_lancode_map(self, translations):
        lancode_map = defaultdict(list)
        for trans in translations:
            lancode_map[trans.langcode].append(trans)

        return lancode_map

    def _get_id_map(self, translations):
        id_map = defaultdict(list)
        for trans in translations:
            id_map[trans.textid].append(trans)

        return id_map

    def update_t2_with_trans(self, t2_xlsx, trans_xlsx):
        translations = self.load_translate_xlsx(trans_xlsx)
        id_map = self._get_id_map(translations)

        wb = load_workbook(filename = t2_xlsx)
        ws = wb.get_sheet_by_name('MESSAGE')
        
        rows = ws.rows
        col_map = {name:index for index, name in enumerate(T2_COL_NAME)}
        for row in rows:
            textid = row[col_map[ID_TITLE]].value
            if textid in id_map:
                trans_of_id = id_map[textid]
                for trans in trans_of_id:
                    #todo:not finished yet
                    pass
                
    def load_langcode(self, langfile='tb_lang.txt'):
        lines = utility.read_as_list(langfile)
        # create code -> name map
        pattern = re.compile(r'(\S+)\s+(\S+)\s+(\S+)')
        #more accurate match for languanges with () in name
        pattern2 = re.compile(r'(\S+)\s+(\S+ \(\S+\))\s+(\S+)')
        name_map = {}
        code_map = {}
        for line in lines:
            match = pattern.search(line)
            if match:
                match2 = pattern2.search(line)
                if match2:
                    match = match2

                name = match.groups()[1]
                code = match.groups()[2]
                name_map[code] = name
                code_map[name] = code

        return name_map, code_map
            
    def load_tb_transview(self, tb_transview):
        lanname_map, code_map = self.load_langcode()
        lines = utility.read_as_list(tb_transview)
        translations = []

        pattern = \
        re.compile(r'(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(.+)')
        pattern2 = \
        re.compile(r'(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(\S+ \(\S+\))\s+(.+)')
        for index, line in enumerate(lines):

            match = pattern.match(line)
            if match:
                match2 = pattern2.match(line)
                if match2:
                    match = match2
                text_id = match.groups()[0]
                account = match.groups()[1]
                project = match.groups()[2]
                feature = match.groups()[3]
                lang = match.groups()[4]
                trans = match.groups()[5]

                if lang not in code_map:
                    print(lang, index)
                trans_item = \
                TranslateItem(account, project, feature, 
                    text_id, code_map[lang], trans)

                translations.append(trans_item)
            else:
                print(index)

        return translations


    def convert_excel_trans(self, trans_xlsx, t2_xlsx): 
        translations = self.load_translate_folder(trans_xlsx)
        self.save_trans_t2(translations, t2_xlsx)

    def convert_txt_trans(self, tb_transview, t2_xlsx):
        translations = self.load_tb_transview(tb_transview)
        self.save_trans_t2(translations, t2_xlsx)

    def convert_trans_bytextids(self, t2_xlsx):
        server = 'BJ-SQL02.fihtdc.com'
        user='nString'
        password='Nstring123456'

        with pymssql.connect(server, user, password, 'nstring',charset='utf8')\
            as conn:
            with conn.cursor(as_dict=True) as cursor:
                textids = self.get_textids()

                sql = "select Text, Account, Project, Feature, Language, Lv1 \
                from TranslationView where Account='KAIOS_UPDATE' and Text in " \
                + textids

                cursor.execute(sql)

                lines = []
                translations = []
                for row in cursor:
                    trans_item = \
                    TranslateItem(row['Account'], row['Project'],
                    row['Feature'],row['Text'], row['Language'],
                    row['Lv1'])

                    translations.append(trans_item)

                self.save_trans_t2_v2(translations, t2_xlsx)

    def convert_trans_byaccount(self, t2_xlsx):
        server = 'BJ-SQL02.fihtdc.com'
        user='nString'
        password='Nstring123456'

        with pymssql.connect(server, user, password, 'nstring',charset='utf8')\
            as conn:
            with conn.cursor(as_dict=True) as cursor:

                sql = "select Text, Account, Project, Feature, Language, Lv1 \
                from TranslationView where Account='S30subcont' and \
                Project='simplex4'"

                cursor.execute(sql)

                lines = []
                translations = []
                for row in cursor:
                    trans_item = \
                    TranslateItem(row['Account'], row['Project'],
                    row['Feature'],row['Text'], row['Language'],
                    row['Lv1'])

                    translations.append(trans_item)

                self.save_trans_t2_v2(translations, t2_xlsx)



    def get_textids(self, fname = 'tb_searched.txt'):
        with open(fname,'r',encoding='utf8') as f:
            textids="("
            for line in f.readlines():
                line = line.strip()
                if line:
                    textids = "{0}'{1}',".format(textids, line)
            textids = textids[:-1]  + ")"#remove last comma

        return textids


    def save_trans_t2_v2(self, translations, t2_xlsx):
        # don't map lancode and langmap in this version
        id_map = self._get_id_map(translations)
        lancode_map = self._get_lancode_map(translations)

        wb = Workbook()
        ws = wb.active

        # create map from langcode to column index in xlsx
        # fill title row 
        cur_row = 1
        cur_col = 1
        ws.cell(row=cur_row, column=cur_col).value = 'RefName' 
        col_map = {'RefName':cur_col}
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = 'ModOP' 
        col_map['ModOP'] = cur_col

        for code in sorted(lancode_map.keys()):
            cur_col += 1
            ws.cell(row=cur_row, column=cur_col).value = code
            col_map[code] = cur_col 

        # fill translation row
        for textid in sorted(id_map.keys()):
            cur_row += 1
            cur_col = col_map['RefName']
            ws.cell(row=cur_row, column=cur_col).value = textid
            cur_col = col_map['ModOP']
            ws.cell(row=cur_row, column=cur_col).value = \
                    id_map[textid][0].feature
            for trans in id_map[textid]:
                cur_col = col_map[trans.langcode]
                ws.cell(row=cur_row, column=cur_col).value = trans.translation

        wb.save(t2_xlsx)



    def save_trans_t2(self, translations, t2_xlsx):
        id_map = self._get_id_map(translations)
        lancode_map = self._get_lancode_map(translations)
        lanname_map, code_map = self.load_langcode()

        wb = Workbook()
        ws = wb.active

        # create map from langcode to column index in xlsx
        # fill title row 
        cur_row = 1
        cur_col = 1
        ws.cell(row=cur_row, column=cur_col).value = 'RefName' 
        col_map = {'RefName':cur_col}
        cur_col += 1
        ws.cell(row=cur_row, column=cur_col).value = 'ModOP' 
        col_map['ModOP'] = cur_col

        for code in sorted(lancode_map.keys()):
            cur_col += 1
            ws.cell(row=cur_row, column=cur_col).value = \
                lanname_map[code]
            col_map[code] = cur_col 

        # fill translation row
        for textid in sorted(id_map.keys()):
            cur_row += 1
            cur_col = col_map['RefName']
            ws.cell(row=cur_row, column=cur_col).value = textid
            cur_col = col_map['ModOP']
            ws.cell(row=cur_row, column=cur_col).value = \
                    id_map[textid][0].feature
            for trans in id_map[textid]:
                cur_col = col_map[trans.langcode]
                ws.cell(row=cur_row, column=cur_col).value = trans.translation

        wb.save(t2_xlsx)

    def trans_to_source(self, source_us, source_xlsx, trans_xlsx):
        """
        for each langcode in trans_xlsx, copy source_us to langcode
            named folder, and update the .properties files according to
            source_xlsx and trans_xlsx
        """
        import copy
        import shutil

        source_map = picksource.SourceToXlsx.load_source_xlsx(source_xlsx)
        translations = self.load_translate_xlsx(trans_xlsx)
        lancode_map = self._get_lancode_map(translations)

        source_root = os.path.abspath(os.path.join(source_us, os.pardir))
        for lancode in lancode_map:
            source_lan = os.path.join(source_root, lancode)

            if not os.path.exists(source_lan):
                shutil.copytree(source_us, source_lan)

            # map source file and all it's translate string
            trans_source_map = defaultdict(list)
            for tran in lancode_map[lancode]:
                src_item = source_map[tran.textid]
                tran_src_item = copy.copy(src_item)
                tran_src_item.value = tran.translation
                trans_source_map[tran_src_item.res_file].append(tran_src_item)

            self.update_source_lan(source_lan, trans_source_map)


    def update_source_lan(self, source_lan, source_map):
        for relpath in source_map:
            picksource.SourceUpdate().update_file(os.path.join(source_lan,
                relpath), source_map[relpath])


def to_t2_xlsx_main():
    parser = argparse.ArgumentParser(description='Create excel file\
            with translation filled in t2 format')

    parser.add_argument('-i','--in', nargs='+', 
                       help='The folder including translation table exported from \
                       nstring')
    parser.add_argument('-o','--out', nargs='+', 
                      help='output file. translation table in t2 \
                      format')
    args = vars(parser.parse_args())
    nstring_trans = 'string_src/translation/'
    t2_trans = 'string_src/trans_t2.xlsx'
    if args['in']:
        nstring_trans = args['in'][0]
    if args['out']:
        t2_trans = args['out'][0]

    convertor = XlsxToSource()
    # t2_trans = 'trans_t2_newids.xlsx'
    # convertor.convert_trans_bytextids(t2_trans)
    t2_trans = 'trans_neo.xlsx'
    convertor.convert_trans_byaccount(t2_trans)

def to_source_file_main():

    parser = argparse.ArgumentParser(description='Create translated \
    source files from base source files.')

    parser.add_argument('-b','--base', nargs='+', required=True,
                       help='''Define root path of resource files.
                       Normally it is the path of en-US folder.''')

    parser.add_argument('-s','--sourcetable', nargs='+', 
                       help='''The source table (xlsx file) generated 
                       from base source files. If not given, it will be
                       set to source.xlsx which has the same parent with
                       path set by --base''')

    parser.add_argument('-t','--translationtable', nargs='+',
                       help='''The translation table (xlsx file) which
                       contains the translated string. The translation
                       will be update to source file according the ID
                       mapping in source table.''')
   
    args = vars(parser.parse_args())
   
    source_base = args['base'][0]
    parent = os.path.abspath(os.path.join(source_base, os.pardir))

    if args['sourcetable']:
        source_table = args['sourcetable'][0]
    else:
        source_table = os.path.join(parent, 'source.xlsx')

    if args['translationtable']:
        trans_table = args['translationtable'][0]
    else:
        trans_table = os.path.join(parent, 'trans.xlsx')

    convertor = XlsxToSource()
    convertor.trans_to_source(source_base, source_table, trans_table)

if __name__ == '__main__':
    to_t2_xlsx_main()


